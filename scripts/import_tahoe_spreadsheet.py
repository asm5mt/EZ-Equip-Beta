import calendar
import datetime as dt
import json
import re
import sqlite3
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data.db"
XLSX_PATH = Path("/home/user/workspace/Auto_Mainten_Tracker_05_Tahoe.xlsx")
TARGET_VIN = "1GNEK13T05R204731"


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def title_case_service(value: str) -> str:
    exceptions = {"ATF", "A/C", "HVAC", "TPMS", "OE", "GM", "DIY"}
    words = re.split(r"(\s+|-|/)", value.title())
    return "".join(w.upper() if w.upper() in exceptions else w for w in words)


def parse_date(value):
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial fallback.
        return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(value))).date()
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y"):
            try:
                return dt.datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    raise ValueError(f"Unrecognized date value: {value!r}")


def ts_seconds(date_value):
    date_value = parse_date(date_value)
    if date_value is None:
        return None
    return calendar.timegm(dt.datetime.combine(date_value, dt.time()).timetuple())


def row_headers(ws, header_row):
    headers = []
    for col in range(1, ws.max_column + 1):
        header = clean(ws.cell(header_row, col).value)
        if header:
            headers.append((col, str(header)))
    return headers


def cell_map(ws, row, headers):
    return {header: clean(ws.cell(row, col).value) for col, header in headers}


def note_join(parts):
    return "\n".join(str(part) for part in parts if part not in (None, ""))


def service_category(title):
    lower = title.lower()
    if any(k in lower for k in ("oil", "air filter", "battery", "belt")):
        return "engine"
    if any(k in lower for k in ("trans", "transfer", "diff")):
        return "drivetrain"
    if "wiper" in lower:
        return "other"
    return "other"


def schedule_name(sheet_title):
    mapping = {
        "OIL CHANGE": "Oil Change",
        "AIR FILTER": "Air Filter",
        "TRANSMISSION": "Transmission Service",
        "TRANSFER CASE": "Transfer Case Service",
        "FRONT DIFFERENTIAL": "Front Differential Service",
        "REAR DIFFERENTIAL": "Rear Differential Service",
        "WIPERS": "Windshield Wipers",
        "BELT": "Serpentine Belt",
        "BATTERY": "Battery Terminal Service",
    }
    return mapping.get(str(sheet_title).upper(), title_case_service(str(sheet_title)))


def insert_line(cur, event_id, item_name, part_number=None, brand=None, spec=None, unit=None, notes=None):
    item_name = clean(item_name)
    if not item_name:
        return
    cur.execute(
        """
        INSERT INTO service_line_items
          (service_event_id, inventory_item_id, item_name, part_number, brand, spec, quantity, unit, unit_cost, notes)
        VALUES (?, NULL, ?, ?, ?, ?, 1, ?, NULL, ?)
        """,
        (event_id, item_name, clean(part_number), clean(brand), clean(spec), clean(unit), clean(notes)),
    )


def add_line_items_for_service(cur, event_id, sheet_name, title, values):
    viscosity = values.get("Viscosity")
    oil = values.get("Oil")
    fluid = values.get("Fluid")
    filter_value = values.get("Filter")
    service = values.get("Service")
    location = values.get("Location")

    if oil:
        insert_line(cur, event_id, "Engine Oil", brand=oil, spec=viscosity, unit="qt")
    if fluid:
        fluid_name = "Transmission Fluid" if "TRANS" in sheet_name else "Differential Fluid" if "DIFF" in sheet_name else "Transfer Case Fluid" if "T-CASE" in sheet_name else "Fluid"
        insert_line(cur, event_id, fluid_name, brand=fluid, spec=viscosity, unit="qt")
    if filter_value:
        if "OIL" in sheet_name:
            item_name = "Oil Filter"
        elif "AIR" in sheet_name:
            item_name = "Air Filter"
        elif "TRANS" in sheet_name:
            item_name = "Transmission Filter"
        else:
            item_name = "Filter"
        insert_line(cur, event_id, item_name, part_number=filter_value)
    if values.get("Wiper Brand") or values.get("Wiper Model"):
        insert_line(
            cur,
            event_id,
            f"Wiper Blade{f' - {location}' if location else ''}",
            part_number=values.get("Wiper Model"),
            brand=values.get("Wiper Brand"),
            notes=service,
        )
    if values.get("Belt Brand") or values.get("Belt Model"):
        insert_line(
            cur,
            event_id,
            f"Belt{f' - {location}' if location else ''}",
            part_number=values.get("Belt Model"),
            brand=values.get("Belt Brand"),
            notes=service,
        )
    if sheet_name == "BATTERY":
        insert_line(cur, event_id, service or "Battery Terminal Service", notes=location)


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    main_ws = wb["MAIN-HOME"]

    asset_name = clean(main_ws["F3"].value) or "2005 Chevrolet Tahoe LT Z71"
    current_meter = float(main_ws["H4"].value)
    meter_as_of = None
    last_log_date = None
    for row in range(22, main_ws.max_row + 1):
        date_value = parse_date(main_ws.cell(row, 3).value) if clean(main_ws.cell(row, 3).value) else None
        odometer = clean(main_ws.cell(row, 4).value)
        if date_value and odometer is not None:
            last_log_date = date_value
    if last_log_date:
        meter_as_of = ts_seconds(last_log_date)

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = OFF")
    cur = conn.cursor()
    asset = cur.execute("SELECT id, fleet_id FROM assets WHERE vin = ?", (TARGET_VIN,)).fetchone()
    if not asset:
        raise RuntimeError(f"No target asset found with VIN {TARGET_VIN}")
    asset_id, fleet_id = asset

    with conn:
        event_ids = [row[0] for row in cur.execute("SELECT id FROM service_events WHERE asset_id = ?", (asset_id,)).fetchall()]
        if event_ids:
            placeholders = ",".join("?" for _ in event_ids)
            cur.execute(f"DELETE FROM service_line_items WHERE service_event_id IN ({placeholders})", event_ids)
        cur.execute("DELETE FROM service_events WHERE asset_id = ?", (asset_id,))
        cur.execute("DELETE FROM meter_readings WHERE asset_id = ?", (asset_id,))
        cur.execute("DELETE FROM maintenance_schedules WHERE asset_id = ?", (asset_id,))

        cur.execute(
            """
            UPDATE assets
            SET friendly_name = ?,
                asset_type = 'Vehicle',
                year = 2005,
                make = 'Chevrolet',
                model = 'Tahoe',
                trim = 'LT Z71',
                vin = ?,
                engine = 'LQ4',
                transmission = '4L60E',
                drivetrain = '4WD/4-Wheel Drive/4x4',
                fuel_type = 'Gasoline',
                displacement_liters = 6.0,
                engine_cylinders = 8,
                engine_configuration = 'V',
                gvwr = 'Class 2E: 6,001 - 7,000 lb (2,722 - 3,175 kg)',
                body_type = 'Sport Utility Vehicle (SUV)/Multi-Purpose Vehicle (MPV)',
                vin_decoded_fields = ?,
                meter_type = 'mileage',
                meter_label = NULL,
                current_meter = ?,
                meter_as_of = ?,
                is_active = 1,
                inactive_reason = NULL,
                status = 'active',
                notes = ?
            WHERE id = ?
            """,
            (
                asset_name,
                TARGET_VIN,
                json.dumps(["year", "make", "model", "engine", "drivetrain", "fuelType", "displacementLiters", "engineCylinders", "engineConfiguration", "gvwr", "bodyType"]),
                current_meter,
                meter_as_of,
                "Daily driver. Seeded from Auto_Mainten_Tracker_05_Tahoe.xlsx. Spreadsheet source VIN differed from the app target VIN; target VIN retained per user instruction.",
                asset_id,
            ),
        )

        meter_count = 0
        for row in range(22, main_ws.max_row + 1):
            date_value = clean(main_ws.cell(row, 3).value)
            odometer = clean(main_ws.cell(row, 4).value)
            notes = clean(main_ws.cell(row, 6).value)
            if date_value and odometer is not None:
                cur.execute(
                    """
                    INSERT INTO meter_readings (asset_id, reading_type, value, reading_date, notes, source)
                    VALUES (?, 'mileage', ?, ?, ?, 'manual')
                    """,
                    (asset_id, float(odometer), ts_seconds(date_value), note_join([notes, f"Imported from spreadsheet MAIN-HOME row {row}"]) if notes else f"Imported from spreadsheet MAIN-HOME row {row}"),
                )
                meter_count += 1

        schedule_ids = {}
        scheduled_event_count = 0
        line_count = 0
        for ws in wb.worksheets:
            if ws.title in ("MAIN-HOME", "REPAIRS"):
                continue
            raw_title = clean(ws["B3"].value) or ws.title
            name = schedule_name(str(raw_title))
            interval = clean(ws["F7"].value)
            unit = str(clean(ws["G7"].value) or "").upper()
            meter_interval = float(interval) if interval and "MI" in unit else None
            day_interval = int(interval) if interval and "DAY" in unit else None
            reading_type = "mileage" if meter_interval else "date"
            details = clean(ws["B4"].value)
            notes = note_join([details, f"Imported from spreadsheet sheet {ws.title}."])
            cur.execute(
                """
                INSERT INTO maintenance_schedules
                  (asset_id, name, category, reading_type, meter_interval, day_interval, meter_due_soon, day_due_soon, notes, active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                """,
                (
                    asset_id,
                    name,
                    service_category(name),
                    reading_type,
                    meter_interval,
                    day_interval,
                    500 if meter_interval else None,
                    30 if day_interval else None,
                    notes,
                ),
            )
            schedule_id = cur.lastrowid
            schedule_ids[ws.title] = schedule_id

            headers = row_headers(ws, 10)
            for row in range(11, min(ws.max_row, 1000) + 1):
                values = cell_map(ws, row, headers)
                date_value = values.get("Date")
                if not date_value:
                    continue
                performed = ts_seconds(date_value)
                meter = clean(values.get("Changed (Mi.)"))
                if meter is not None:
                    meter = float(meter)
                detail_parts = []
                for key in ("Due (Mi.)", "Due", "Distance (Mi.)", "Location", "Service", "Viscosity", "Oil", "Fluid", "Filter", "Wiper Brand", "Wiper Model", "Belt Brand", "Belt Model", "Notes"):
                    value = clean(values.get(key))
                    if value is not None:
                        detail_parts.append(f"{key}: {value}")
                detail_parts.append(f"Imported from spreadsheet {ws.title} row {row}.")
                cur.execute(
                    """
                    INSERT INTO service_events
                      (asset_id, schedule_id, event_type, title, performed_at, meter_at_service, vendor, technician, cost, notes)
                    VALUES (?, ?, 'scheduled', ?, ?, ?, NULL, NULL, NULL, ?)
                    """,
                    (asset_id, schedule_id, name, performed, meter, note_join(detail_parts)),
                )
                event_id = cur.lastrowid
                before = cur.execute("SELECT COUNT(*) FROM service_line_items").fetchone()[0]
                add_line_items_for_service(cur, event_id, ws.title, name, values)
                after = cur.execute("SELECT COUNT(*) FROM service_line_items").fetchone()[0]
                line_count += after - before
                scheduled_event_count += 1

        repairs = wb["REPAIRS"]
        repair_count = 0
        for row in range(8, repairs.max_row + 1):
            date_value = clean(repairs.cell(row, 2).value)
            if not date_value:
                continue
            odometer = clean(repairs.cell(row, 4).value)
            facility = clean(repairs.cell(row, 6).value)
            service = clean(repairs.cell(row, 8).value) or "Repair"
            notes = clean(repairs.cell(row, 10).value)
            cur.execute(
                """
                INSERT INTO service_events
                  (asset_id, schedule_id, event_type, title, performed_at, meter_at_service, vendor, technician, cost, notes)
                VALUES (?, NULL, 'repair', ?, ?, ?, ?, NULL, NULL, ?)
                """,
                (
                    asset_id,
                    title_case_service(str(service)),
                    ts_seconds(date_value),
                    float(odometer) if odometer is not None else None,
                    facility,
                    note_join([notes, f"Imported from spreadsheet REPAIRS row {row}."]),
                ),
            )
            event_id = cur.lastrowid
            insert_line(cur, event_id, title_case_service(str(service)), notes=notes)
            line_count += 1
            repair_count += 1

    summary = {
        "asset_id": asset_id,
        "asset_name": asset_name,
        "target_vin": TARGET_VIN,
        "current_meter": current_meter,
        "meter_as_of": last_log_date.isoformat() if last_log_date else None,
        "meter_readings_imported": meter_count,
        "schedules_imported": len(schedule_ids),
        "scheduled_service_events_imported": scheduled_event_count,
        "repair_events_imported": repair_count,
        "service_line_items_imported": line_count,
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
